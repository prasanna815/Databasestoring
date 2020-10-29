from openpyxl import *
from tkinter import *
wb=load_workbook(r'C:/Users/91812/Desktop/data.xlsx')
sheet=wb.active
def excel():
    sheet.column_dimensions['A'].width=30
    sheet.column_dimensions['B'].width=30
    sheet.column_dimensions['C'].width=50
    sheet.cell(row=1,column=1).value="NAME"
    sheet.cell(row=1,column=2).value="PHONE"
    sheet.cell(row=1,column=3).value="EMAIL"
def clear():
    name_field.delete(0,END)
    phone_field.delete(0,END)
    email_field.delete(0,END)
def insert():
    if(name_field.get()=="" and phone_field.get()=="" and email_field.get()==""):
        print("Please enter all fields")
    else:
        current_row=sheet.max_row
        current_column=sheet.max_column
        sheet.cell(row=current_row+1,column=1).value=name_field.get()
        sheet.cell(row=current_row+1,column=2).value=phone_field.get()
        sheet.cell(row=current_row+1,column=3).value=email_field.get()
        wb.save(r'C:/Users/91812/Desktop/data.xlsx')
        clear()
if __name__=="__main__":
    root=Tk()
    root.title("Register Here")
    excel()
    name=Label(root,text="Name")
    name.grid(row=1,column=0)
    phone=Label(root,text="Phone")
    phone.grid(row=2,column=0)
    email=Label(root,text="Email")
    email.grid(row=3,column=0)
    name_field=Entry(root)
    name_field.grid(row=1,column=1,ipadx="100")
    phone_field=Entry(root)
    phone_field.grid(row=2,column=1,ipadx="100")
    email_field=Entry(root)
    email_field.grid(row=3,column=1,ipadx="100")
    excel()
    submit=Button(root,text="Submit",command=insert)
    submit.grid(row=8,column=1)
    root.mainloop()

    
        
        
    
